from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, Border, Side


def create_excel(df):
    buf = BytesIO()

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule')
        wb = writer.book
        ws = writer.sheets['Schedule']
        # 全てのセルのフォントを変更
        font = Font(name='HG丸ｺﾞｼｯｸM-PRO', size=10)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

        # 2列目以降の列幅を3.30に設定（40ピクセルに相当）
        for col in ws.columns:
            if col[0].column == 1:
                ws.column_dimensions[col[0].column_letter].width = 17.60

            else:  # 2列目以降
                ws.column_dimensions[col[0].column_letter].width = 4.40

        for i in range(3, ws.max_row + 1):  # 1行目をスキップ
            if i % 2 == 0:  # 偶数行
                ws.row_dimensions[i].height = 10.8
            else:  # 奇数行
                ws.row_dimensions[i].height = 18.2

        # 境界線のスタイルを定義
        thin = Side(border_style="thin", color="000000")
        dotted = Side(border_style="dotted", color="000000")

        # 全セルに格子状の線を適用
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row, values_only=False):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.border = Border(
                        left=thin, right=thin, bottom=thin, top=thin if cell.row == 2 else None)
                else:
                    cell.border = Border(
                        top=thin, left=thin, right=thin, bottom=dotted)

    buf.seek(0)

    return buf
